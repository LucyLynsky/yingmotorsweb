# -*- coding: utf-8 -*-
"""Build Yingmotors product/media Excel datasheets from js/products.js + images/stock."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

ROOT = Path(r"E:\codePrj\web")
JS = (ROOT / "js" / "products.js").read_text(encoding="utf-8")
STOCK = ROOT / "images" / "stock"
OUT = ROOT / "docs" / "Yingmotors-product-media-datasheet.xlsx"

TYPE_META = {
    "truck": ("重卡", "Trucks", "02_重卡_Trucks"),
    "trailer": ("挂车", "Trailers", "03_挂车_Trailers"),
    "tricycle": ("三轮车", "Tricycles", "04_三轮四轮_Light"),
    "fourwheel": ("四轮 / UTV", "UTV / 4WD", "04_三轮四轮_Light"),
    "bus": ("客车", "Buses", "05_客车_Buses"),
    "excavator": ("挖掘机", "Excavators", "06_挖掘机_Excavators"),
    "loader": ("装载机", "Loaders", "09_装载机_Loaders"),
    "mixer": ("搅拌车", "Mixers", "07_搅拌车_Mixers"),
    "special": ("环卫车", "Sanitation", "08_环卫车_Sanitation"),
}

GOLD = "E49402"
INK = "210E06"
PAPER = "F6F1EA"
WHITE = "FFFFFF"
ROW_ALT = "FFF8F0"

thin = Border(
    left=Side(style="thin", color="D9C9B3"),
    right=Side(style="thin", color="D9C9B3"),
    top=Side(style="thin", color="D9C9B3"),
    bottom=Side(style="thin", color="D9C9B3"),
)
head_fill = PatternFill("solid", fgColor=INK)
head_font = Font(name="Microsoft YaHei", bold=True, color=GOLD, size=11)
body_font = Font(name="Microsoft YaHei", size=10, color=INK)
title_font = Font(name="Microsoft YaHei", bold=True, size=16, color=INK)
wrap = Alignment(wrap_text=True, vertical="center")
left = Alignment(wrap_text=True, vertical="center", horizontal="left")


def ym_stock(sku: str, slug: str, count: int, thumb_no: int = 1) -> dict:
    folder = f"{sku}_{slug}"
    images = [f"images/stock/{folder}/{sku}_{i:02d}.jpg" for i in range(1, count + 1)]
    thumb = f"images/stock/{folder}/thumbs/{sku}_{thumb_no:02d}.jpg"
    return {"images": images, "thumb": thumb}


def parse_helpers(src: str) -> dict:
    helpers = {}
    for m in re.finditer(
        r'var _(\w+) = ymStock\("([^"]+)",\s*"([^"]+)",\s*(\d+)(?:,\s*(\d+))?\)',
        src,
    ):
        key, sku, slug, count, thumb = m.group(1), m.group(2), m.group(3), int(m.group(4)), m.group(5)
        helpers["_" + key] = ym_stock(sku, slug, count, int(thumb) if thumb else 1)
    return helpers


def extract_block_field(block: str, field: str) -> str | None:
    m = re.search(rf"\b{field}:\s*\"([^\"]*)\"", block)
    return m.group(1) if m else None


def extract_lang_name(block: str, lang: str, key: str) -> str:
    m = re.search(rf"{lang}:\s*\{{(.*?)\n    \}}", block, re.S)
    if not m:
        return ""
    inner = m.group(1)
    n = re.search(rf"{key}:\s*\"([^\"]*)\"", inner)
    return n.group(1) if n else ""


def resolve_images(block: str, helpers: dict) -> list[str]:
    m = re.search(r"images:\s*(.*?)\n\s*thumb:", block, re.S)
    if not m:
        return []
    expr = m.group(1)
    out = []
    for hm in re.finditer(r"(_[a-z0-9]+)\.images(?:\[(\d+)\])?", expr):
        imgs = helpers[hm.group(1)]["images"]
        if hm.group(2) is not None:
            i = int(hm.group(2))
            if i < len(imgs):
                out.append(imgs[i])
        else:
            out.extend(imgs)
    if out:
        return out
    return re.findall(r"\"(images/stock/[^\"]+)\"", expr)


def resolve_thumb(expr: str, helpers: dict) -> str:
    expr = expr.strip().rstrip(",")
    if expr.endswith(".thumb"):
        return helpers[expr.split(".")[0]]["thumb"]
    m = re.search(r"\"(images/stock/[^\"]+)\"", expr)
    return m.group(1) if m else ""


def resolve_videos(block: str) -> list[str]:
    out = []
    for m in re.finditer(
        r'ymVid\("([^"]+)",\s*"([^"]+)",\s*\[([^\]]*)\]\)',
        block,
    ):
        sku, slug = m.group(1), m.group(2)
        names = re.findall(r'"([^"]+)"', m.group(3))
        folder = f"{sku}_{slug}"
        for n in names:
            name = n if n.startswith(sku + "_") else f"{sku}_{n}"
            out.append(f"images/stock/{folder}/{name}")
    return out


def parse_products(src: str, helpers: dict) -> list[dict]:
    body = src.split("window.YM_PRODUCTS = [", 1)[1].split("];", 1)[0]
    chunks = re.split(r"\n  \{\n", body)
    products = []
    for chunk in chunks:
        if "id:" not in chunk:
            continue
        block = "  {\n" + chunk
        pid = extract_block_field(block, "id")
        if not pid:
            continue
        th_m = re.search(r"thumb:\s*(.+?)\n", block)
        images = resolve_images(block, helpers)
        thumb = resolve_thumb(th_m.group(1), helpers) if th_m else ""
        ptype = extract_block_field(block, "type") or ""
        tz, te, sheet = TYPE_META[ptype]
        products.append(
            {
                "id": pid,
                "sku": extract_block_field(block, "sku") or "",
                "category": extract_block_field(block, "category") or "",
                "type": ptype,
                "type_zh": tz,
                "type_en": te,
                "sheet": sheet,
                "brand": extract_block_field(block, "brand") or "",
                "name_zh": extract_lang_name(block, "zh", "name"),
                "name_en": extract_lang_name(block, "en", "name"),
                "subtitle_zh": extract_lang_name(block, "zh", "subtitle"),
                "subtitle_en": extract_lang_name(block, "en", "name") and extract_lang_name(block, "en", "subtitle"),
                "summary_zh": extract_lang_name(block, "zh", "summary"),
                "summary_en": extract_lang_name(block, "en", "summary"),
                "images": images,
                "thumb": thumb,
                "videos": resolve_videos(block),
            }
        )
    # fix subtitle_en (I mixed the condition)
    for p in products:
        # re-parse subtitle_en properly from original — already used extract_lang_name for subtitle in en
        pass
    # featured
    feat = re.search(r"window\.YM_FEATURED_IDS = \[([^\]]+)\]", src)
    featured = set(re.findall(r"\"([^\"]+)\"", feat.group(1))) if feat else set()
    for p in products:
        p["featured"] = "Y" if p["id"] in featured else "N"
        p["category_zh"] = "新车" if p["category"] == "new" else "二手"
        p["category_en"] = "New" if p["category"] == "new" else "Used"
        p["detail_url"] = f"product.html?id={p['id']}"
        p["filter_url"] = f"products.html?cat={p['type']}"
        if p["type"] in ("tricycle", "fourwheel"):
            p["filter_url"] = "products.html?cat=light"
        if p["type"] in ("excavator", "loader", "mixer", "special"):
            p["filter_url"] = "products.html?cat=machinery"
    return products


def file_info(rel: str) -> dict:
    path = ROOT / rel.replace("/", "\\")
    info = {
        "exists": "Y" if path.is_file() else "N",
        "size_bytes": 0,
        "size_kb": "",
        "width": "",
        "height": "",
        "abs_path": str(path),
        "filename": path.name,
        "ext": path.suffix.lower(),
    }
    if not path.is_file():
        return info
    info["size_bytes"] = path.stat().st_size
    info["size_kb"] = round(path.stat().st_size / 1024, 1)
    if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
        try:
            with Image.open(path) as im:
                info["width"], info["height"] = im.size
        except Exception:
            pass
    return info


def style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin
    ws.row_dimensions[row].height = 28
    ws.auto_filter.ref = f"A{row}:{get_column_letter(cols)}{row}"
    ws.freeze_panes = f"A{row + 1}"


def write_rows(ws, headers: list[str], rows: list[list], start_row: int = 1):
    for i, h in enumerate(headers, 1):
        ws.cell(start_row, i, h)
    style_header(ws, start_row, len(headers))
    for r_i, row in enumerate(rows, start_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if (r_i - start_row) % 2 else ROW_ALT)
        for c, val in enumerate(row, 1):
            cell = ws.cell(r_i, c, val)
            cell.font = body_font
            cell.alignment = wrap
            cell.border = thin
            cell.fill = fill
        ws.row_dimensions[r_i].height = 22
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    return start_row + len(rows)


def set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def product_headers():
    return [
        "sku 现车编号",
        "product_id",
        "状态_zh",
        "condition_en",
        "分类_zh",
        "type_en",
        "品牌 brand",
        "中文名称",
        "English name",
        "副标题_zh",
        "subtitle_en",
        "简介_zh",
        "summary_en",
        "首页推荐",
        "详情页",
        "列表筛选",
        "封面图 thumb",
        "图片数",
        "视频数",
        "图片路径 images",
        "视频路径 videos",
    ]


def product_row(p: dict) -> list:
    return [
        p.get("sku") or "",
        p["id"],
        p["category_zh"],
        p["category_en"],
        p["type_zh"],
        p["type_en"],
        p["brand"],
        p["name_zh"],
        p["name_en"],
        p["subtitle_zh"],
        p["subtitle_en"],
        p["summary_zh"],
        p["summary_en"],
        p["featured"],
        p["detail_url"],
        p["filter_url"],
        p["thumb"],
        len(p["images"]),
        len(p["videos"]),
        " | ".join(p["images"]),
        " | ".join(p["videos"]) if p["videos"] else "",
    ]


PROD_WIDTHS = [14, 22, 10, 12, 12, 14, 14, 28, 36, 28, 32, 40, 40, 10, 28, 24, 36, 10, 10, 55, 45]


def collect_media(products: list[dict]) -> list[dict]:
    rows = []
    for p in products:
        seq = 0
        # thumb first
        seq += 1
        info = file_info(p["thumb"])
        rows.append(
            {
                "sku": p.get("sku") or "",
                "product_id": p["id"],
                "name_zh": p["name_zh"],
                "name_en": p["name_en"],
                "category_zh": p["category_zh"],
                "type_zh": p["type_zh"],
                "role": "thumb",
                "role_zh": "封面缩略图",
                "seq": seq,
                "web_path": p["thumb"],
                **info,
            }
        )
        for i, rel in enumerate(p["images"], 1):
            seq += 1
            info = file_info(rel)
            rows.append(
                {
                    "sku": p.get("sku") or "",
                    "product_id": p["id"],
                    "name_zh": p["name_zh"],
                    "name_en": p["name_en"],
                    "category_zh": p["category_zh"],
                    "type_zh": p["type_zh"],
                    "role": "gallery",
                    "role_zh": f"详情图 {i}",
                    "seq": seq,
                    "web_path": rel,
                    **info,
                }
            )
        for i, rel in enumerate(p["videos"], 1):
            seq += 1
            info = file_info(rel)
            rows.append(
                {
                    "sku": p.get("sku") or "",
                    "product_id": p["id"],
                    "name_zh": p["name_zh"],
                    "name_en": p["name_en"],
                    "category_zh": p["category_zh"],
                    "type_zh": p["type_zh"],
                    "role": "video",
                    "role_zh": f"详情视频 {i}",
                    "seq": seq,
                    "web_path": rel,
                    **info,
                }
            )
    return rows


def unlinked_stock(products: list[dict]) -> list[list]:
    used = set()
    for p in products:
        used.add(p["thumb"].replace("\\", "/"))
        used.update(x.replace("\\", "/") for x in p["images"])
        used.update(x.replace("\\", "/") for x in p["videos"])
    rows = []
    for p in sorted(STOCK.rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in {".jpg", ".jpeg", ".png", ".mp4", ".webm"}:
            continue
        rel = p.relative_to(ROOT).as_posix()
        if rel in used:
            continue
        kind = "video" if p.suffix.lower() == ".mp4" else "image"
        info = file_info(rel)
        rows.append(
            [
                rel,
                kind,
                info["filename"],
                info["size_kb"],
                info["width"],
                info["height"],
                info["abs_path"],
                "网站 YM_PRODUCTS 未引用，磁盘上仍有文件",
            ]
        )
    return rows


def fill_readme(ws, products, media):
    ws["A1"] = "YING MOTORS 产品媒体资料表"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    lines = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("用途", "后期数据库维护依据：产品主数据 + 网站引用的图片/视频路径"),
        ("数据来源", "js/products.js（网站现车） + images/stock/（网页实际引用文件）"),
        ("产品条数", len(products)),
        ("媒体文件行数（含封面+详情图+视频）", len(media)),
        ("缺文件数", sum(1 for m in media if m["exists"] != "Y")),
        ("", ""),
        ("工作表说明", ""),
        ("01_产品主表", "一条产品一行。sku 为对外现车编号；product_id 为网站 URL 主键。"),
        ("02～08 分类表", "与网站筛选一致：重卡 / 挂车 / 三轮四轮 / 客车 / 挖掘机 / 搅拌车 / 环卫车。三轮与四轮合在同一张表。"),
        ("09_媒体文件", "网站引用的每一张图、每一个视频一行。role=thumb/gallery/video。"),
        ("10_未引用文件", "images/stock 中存在、但当前网站未挂到任何产品的文件（例如多余视频）。"),
        ("", ""),
        ("字段约定", ""),
        ("sku", "对外现车编号 YM-{N|U}{TYPE}-{NNN}。N=新车 U=二手；TK重卡 TL挂车 TC三轮 FW四轮 BS客车 EX挖掘机 LD装载机 MX搅拌车 SP环卫。同类序号不复用。"),
        ("product_id", "网站 URL：product.html?id=该字段，也可用 sku 打开同一页"),
        ("状态", "new=新车 / used=二手"),
        ("分类", "对应 js 里的 type，以及 products.html?cat= 筛选"),
        ("web_path", "网站相对路径，部署后保持此相对位置"),
        ("exists", "Y=磁盘上有文件，N=网站引用了但文件缺失"),
        ("", ""),
        ("维护建议", "改现车时：先改本表或数据库，再同步 js/products.js 与 images/stock。重新生成：python docs/build_product_datasheet.py"),
    ]
    for i, (k, v) in enumerate(lines, 3):
        ws.cell(i, 1, k).font = Font(name="Microsoft YaHei", bold=True, size=10, color=INK)
        ws.cell(i, 2, v).font = body_font
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(i, 2).alignment = wrap
        ws.row_dimensions[i].height = 20
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 88


def main():
    helpers = parse_helpers(JS)
    products = parse_products(JS, helpers)
    # repair subtitle_en
    body = JS.split("window.YM_PRODUCTS = [", 1)[1].split("];", 1)[0]
    chunks = re.split(r"\n  \{\n", body)
    by_id = {p["id"]: p for p in products}
    for chunk in chunks:
        if "id:" not in chunk:
            continue
        block = "  {\n" + chunk
        pid = extract_block_field(block, "id")
        if pid in by_id:
            by_id[pid]["subtitle_en"] = extract_lang_name(block, "en", "subtitle")

    media = collect_media(products)
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "00_说明"
    fill_readme(ws0, products, media)

    ws1 = wb.create_sheet("01_产品主表")
    write_rows(ws1, product_headers(), [product_row(p) for p in products])
    set_widths(ws1, PROD_WIDTHS)

    # category sheets — keep type order, merge tricycle+fourwheel
    sheet_order = [
        "02_重卡_Trucks",
        "03_挂车_Trailers",
        "04_三轮四轮_Light",
        "05_客车_Buses",
        "06_挖掘机_Excavators",
        "07_搅拌车_Mixers",
        "08_环卫车_Sanitation",
        "09_装载机_Loaders",
    ]
    grouped: dict[str, list] = {s: [] for s in sheet_order}
    for p in products:
        grouped[p["sheet"]].append(p)
    for name in sheet_order:
        ws = wb.create_sheet(name)
        write_rows(ws, product_headers(), [product_row(p) for p in grouped[name]])
        set_widths(ws, PROD_WIDTHS)

    media_headers = [
        "sku 现车编号",
        "product_id",
        "中文名称",
        "English name",
        "状态",
        "分类",
        "用途 role",
        "用途说明",
        "序号",
        "网站路径 web_path",
        "文件名",
        "扩展名",
        "存在",
        "大小KB",
        "宽px",
        "高px",
        "本地绝对路径",
    ]
    media_rows = [
        [
            m.get("sku") or "",
            m["product_id"],
            m["name_zh"],
            m["name_en"],
            m["category_zh"],
            m["type_zh"],
            m["role"],
            m["role_zh"],
            m["seq"],
            m["web_path"],
            m["filename"],
            m["ext"],
            m["exists"],
            m["size_kb"],
            m["width"],
            m["height"],
            m["abs_path"],
        ]
        for m in media
    ]
    ws_m = wb.create_sheet("09_媒体文件_Media")
    write_rows(ws_m, media_headers, media_rows)
    set_widths(ws_m, [14, 22, 28, 36, 10, 12, 12, 14, 8, 48, 18, 10, 8, 12, 10, 10, 70])

    ws_u = wb.create_sheet("10_未引用_Unlinked")
    un = unlinked_stock(products)
    write_rows(
        ws_u,
        ["网站相对路径", "类型", "文件名", "大小KB", "宽px", "高px", "本地绝对路径", "说明"],
        un,
    )
    set_widths(ws_u, [48, 10, 22, 12, 10, 10, 70, 46])

    wb.save(OUT)
    cn = ROOT / "docs" / "Yingmotors-产品媒体资料.xlsx"
    if cn.exists():
        cn.unlink()
    import shutil
    shutil.copy2(OUT, cn)
    missing = [m for m in media if m["exists"] != "Y"]
    print("saved", OUT)
    print("saved", cn)
    print("products", len(products), "media", len(media), "missing", len(missing), "unlinked", len(un))
    for p in products:
        print(f"{p.get('sku','')}\t{p['id']}\t{p['name_zh']}\timg={len(p['images'])}\tvid={len(p['videos'])}")


if __name__ == "__main__":
    main()
